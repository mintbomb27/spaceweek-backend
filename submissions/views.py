import re
from uuid import uuid4
from spaceweekBackend.helpers import GenericResponse
from rest_framework import generics
from .serializers import *
from .models import Event
from .models import Participant
from django.utils import timezone
from rest_framework.permissions import *
from django.shortcuts import render, redirect
from django.shortcuts import render
from django.conf import settings
from django.core.files.storage import default_storage
from django.contrib import messages
import pyrebase
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.response import Response
from openpyxl import *

firebase = pyrebase.initialize_app(settings.FIREBASE_CONFIG)
storage = firebase.storage()

class AllEventsView(generics.ListAPIView):
    serializer_class = EventSerializer
    permission_classes = [IsAuthenticated]
    queryset = Event.objects.all()

    def get(self, request, *args, **kwargs):
        response = super().list(self, request, *args, **kwargs)
        return GenericResponse("success",response.data)

class DetailView(generics.ListAPIView):
    serializer_class = EventSerializer
    permission_classes = [IsAuthenticated]
    queryset = Event.objects.all()

    def get(self, request, *args, **kwargs):
        id = request.GET.get('id', None)
        if(id is None):
            raise Exception(400, 'id not passed')
        event = Event.objects.get(id=id)
        school = School.objects.get(poc=request.user)
        participants = ParticipantSerializer(Participant.objects.filter(event=event, school=school), many=True).data
        for participant in participants:
            participant['team_name']=Team.objects.get(participants__id=participant['id']).name
        return GenericResponse("success", {"event":EventSerializer(event).data, "participants":participants})

class CreateEventView(generics.CreateAPIView):
    serializer_class = EventSerializer
    permission_classes = [IsAdminUser]

    def post(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return GenericResponse("success",response.data)

class DashboardView(generics.GenericAPIView):
    serializer_class = EventDashSerializer

    def get(self, request, *args, **kwargs):
        response = {
            'name':request.user.name,
            'events':EventDashSerializer(Event.objects.all().order_by('id'),many=True).data
        }
        return GenericResponse('success',response)

class ParticipantView(generics.GenericAPIView):
    serializer_class = ParticipantSerializer
    permission_classes = [IsAuthenticated]

    def delete(self, request, *args, **kwargs):
        event_id = request.GET.get('event_id', None)
        if(event_id is None):
            raise Exception(400, 'event_id not passed')
        event = Event.objects.get(id=event_id)
        participant_id = request.GET.get('participant_id', None)
        if(participant_id is None):
            raise Exception(400, 'participant_id not passed')
        participant = Participant.objects.get(id=participant_id)
        school = School.objects.get(poc=request.user)
        if(participant.school != school):
            raise Exception(401, "Participant does not belong to your school.")
        if(participant.event != event):
            raise Exception(400, "Participant is not registered for the mentioned event.")
        part = participant.delete()
        return GenericResponse('Removed Participant from Event',part)

    def post(self, request, *args, **kwargs):
        data = self.get_serializer(data=request.data)
        data.is_valid(raise_exception=True)
        name=data.validated_data.get('name', None)
        if name and name.strip():
            pass
        else:
            raise Exception(400,"Name should not be empty")
        gender=data.validated_data.get('gender', None)
        if gender and gender.strip():
            pass
        else:
            raise Exception(400,"gender should not be empty")
        standard = data.validated_data.get('standard', None)
        if standard and standard.strip():
            pass
        else:
            raise Exception(400,"Strandard should not be empty")
        event_id = request.data.get('event_id', None)
        if(event_id is None):
            raise Exception(400, 'event id not provided')
        event = Event.objects.get(id=event_id)
        if event.deadline: # Check Deadline
            if timezone.now() > event.deadline:
                raise Exception(422, "deadline passed for registration.")
        school = School.objects.get(poc=request.user)
        participants_for_events = Participant.objects.filter(event=event,school=school).count()
        if(standard not in event.eligibility):
            raise Exception(400, "selected standard not eligible for the event.")
        if(participants_for_events >= event.max_per_school):# check if max reached
            raise Exception(400, "maximum participants for events reached")
        if event.deadline: # Check Deadline
            if timezone.now() > event.deadline:
                raise Exception(422, "deadline passed for submission.")
        participant = Participant(name=name,standard=standard, gender=gender, school=school, event=event)
        participant.save()
        return GenericResponse('Registered Student', ParticipantSerializer(participant).data)

class TeamView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TeamSerializer

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file', None)
        event_id = request.data.get('event_id', None)
        if event_id is None:
            raise Exception(400, "event_id not passed")
        event = Event.objects.get(id=event_id)
        if(event.type != 'Team'):
            raise Exception(400, "Mentioned event is not team event.")
        school = School.objects.get(poc=request.user)
        if file is None:
                raise Exception(400, "file not passed")
        participants_for_events = Participant.objects.filter(event=event,school=school).count()
        if(participants_for_events >= event.max_per_school):# check if max reached
            raise Exception(400, "Maximum participants from your school for the event reached")
        ext = file.name.split('.')[-1]
        if(ext in ['xls','xlsx']):
            filename = f"{uuid4()}.{ext}"
            dirr = default_storage.save(filename, file)
            wb = load_workbook(dirr)
            participants = []
            final_parts = []
            teams = {}
            final_teams = []
            curr_teams = Team.objects.all()
            for sheet in wb:
                for row in sheet.iter_rows(min_row=1, max_col=4, max_row=2000, values_only=True):
                    if not row[0]:
                        break
                    if str(row[0]) == 'Name':
                        continue
                    name = row[0]
                    gender = row[1]
                    standard = row[2]
                    team_name = row[3]
                    if(gender.lower() not in ['male','female','other']):
                        delete = default_storage.delete(dirr)
                        raise Exception(400, f'Gender not given as per template. Given {gender}')
                    if(type(standard) is float or type(standard) is int):
                        if(standard < 1 or standard > 12):
                            delete = default_storage.delete(dirr)
                            raise Exception(400, f'`{standard}`: standard not as per given template')
                        else:
                            standard = str(int(standard))
                            if(standard not in event.eligibility):
                                delete = default_storage.delete(dirr)
                                raise Exception(400, f"`{standard}` standard not eligible for the event.")
                    elif(type(standard) is str and standard.lower() != 'college'):
                        delete = default_storage.delete(dirr)
                        raise Exception(400, f'`{standard}`: standard not as per given template')
                    participant = Participant(name=name, gender=gender, standard=standard, event=event, school=school)
                    if(teams.get(team_name, None) is None):
                        teams[team_name] = [participant]
                    else:
                        teams[team_name].append(participant)
            for team_name in teams.keys():
                try:
                    team = curr_teams.get(name=team_name)
                    #team name already exists
                    if(team.school == school):
                        if(team.participants.count() < event.max_per_team):#Team under the same school, and space left
                            if((event.max_per_team - team.participants.count()) >= len(teams[team_name])):#all participants mentioned can enter
                                for part in teams[team_name]:
                                    part.save()
                                    team.participants.add(part)
                                team.save()
                            else:
                                delete = default_storage.delete(dirr)
                                raise Exception(400, f"Team already has {len(team.participants)}. Cannot add {len(teams[team_name])} more.")
                        else:
                            delete = default_storage.delete(dirr)
                            raise Exception(400, f"Team {team.name} already has {event.max_per_team} members")
                    else:
                        delete = default_storage.delete(dirr)
                        raise Exception(400, f"Team already exists under a different school.")
                except Team.DoesNotExist:
                    final_parts += teams[team_name]
                    if(len(teams[team_name]) > event.max_per_team):
                        delete = default_storage.delete(dirr)
                        raise Exception(400, f"Team `{team_name}` has more than {event.max_per_team} participants.")
                    team = Team(name=team_name, event=event, school=school)
                    team.save()
                    for part in teams[team_name]:
                        part.save()
                        team.participants.add(part)
                    team.save()
            # Already registered count
            if(participants_for_events+len(participants) > event.max_per_school):# check if max reached
                delete = default_storage.delete(dirr)
                raise Exception(400, f"You may only add {event.max_per_school-participants_for_events} more participants for the event. Requested to add {len(participants)}.")
            delete = default_storage.delete(dirr)
            return GenericResponse("Registered participants", "Success")
        else:
            raise Exception(400, "Invalid file")

class ParticipantsView(generics.GenericAPIView):
    serializer_class = ParticipantSerializer
    permission_classes = [IsAuthenticated]

    def delete(self, request, *args, **kwargs):
        event_id = request.GET.get('event_id', None)
        if event_id is None:
            raise Exception(400, "event_id not passed")
        event = Event.objects.get(id=event_id)
        school = School.objects.get(poc=request.user)
        Participant.objects.filter(event=event, school=school).delete()
        return GenericResponse('delete all participants for the event', '')

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file', None)
        event_id = request.data.get('event_id', None)
        if event_id is None:
            raise Exception(400, "event_id not passed")
        event = Event.objects.get(id=event_id)
        school = School.objects.get(poc=request.user)
        if file is None:
                raise Exception(400, "file not passed")
        participants_for_events = Participant.objects.filter(event=event,school=school).count()
        if(participants_for_events >= event.max_per_school):# check if max reached
            raise Exception(400, "Maximum participants from your school for the event reached")
        ext = file.name.split('.')[-1]
        if(ext in ['xls','xlsx']):
            filename = f"{uuid4()}.{ext}"
            dirr = default_storage.save(filename, file)
            wb = load_workbook(dirr)
            participants = []
            for sheet in wb:
                for row in sheet.iter_rows(min_row=1, max_col=3, max_row=2000, values_only=True):
                    if not row[0]:
                        break
                    if str(row[0]) == 'Name':
                        continue
                    name = row[0]
                    gender = row[1]
                    standard = row[2]
                    if(gender.lower() not in ['male','female','other']):
                        delete = default_storage.delete(dirr)
                        raise Exception(400, f'Gender not given as per template. Given {gender}')
                    if(type(standard) is float or type(standard) is int):
                        if(standard < 1 or standard > 12):
                            delete = default_storage.delete(dirr)
                            raise Exception(400, f'`{standard}`: standard not as per given template')
                        else:
                            standard = str(int(standard))
                            if(standard not in event.eligibility):
                                delete = default_storage.delete(dirr)
                                raise Exception(400, f"`{standard}` standard not eligible for the event.")
                    elif(type(standard) is str and standard.lower() != 'college'):
                        delete = default_storage.delete(dirr)
                        raise Exception(400, f'`{standard}`: standard not as per given template')
                    participant = Participant(name=name, gender=gender, standard=standard, event=event, school=school)
                    participants.append(participant)
            # Already registered count
            if(participants_for_events+len(participants) > event.max_per_school):# check if max reached
                delete = default_storage.delete(dirr)
                raise Exception(400, f"You may only add {event.max_per_school-participants_for_events} more participants for the event. Requested to add {len(participants)}.")
            Participant.objects.bulk_create(participants)#ignore_conflicts
            delete = default_storage.delete(dirr)
            return GenericResponse("registered participants", "Success")
        else:
            raise Exception(400, "Invalid file")